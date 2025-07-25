import openpyxl
from pathlib import Path
from functions import calculate_date
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from api_solution import request_data


def main():
    
    get_and_save_to_file(Path('./save_test_result.xlsx'))

def get_and_save_to_file(path: Path):
    date_start, date_end = calculate_date()
    wb = openpyxl.Workbook()

    sheet = wb.active

    create_header(sheet)
    usd_data = request_data(date_start.isoformat(), date_end.isoformat(), curr="USD")
    jpy_data = request_data(date_start.isoformat(), date_end.isoformat(), curr="JPY")
    
    max_row_num = fill_data(sheet, usd_data, jpy_data)
    apply_format(sheet, max_row_num)

    wb.save(path)

    return max_row_num - 1

def save_to_file(path: Path, usd_data, jpy_data):
    wb = openpyxl.Workbook()

    sheet = wb.active

    create_header(sheet)
    max_row_num = fill_data(sheet, usd_data=usd_data, jpy_data=jpy_data)
    apply_format(sheet, max_row_num)

    wb.save(path)

    return max_row_num - 1

def create_header(sheet):
    sheet['A1'] = 'Дата USD/RUB'
    sheet['B1'] = 'Курс USD/RUB'
    sheet['C1'] = 'Время USD/RUB'

    sheet['D1'] = 'Дата JPY/RUB'
    sheet['E1'] = 'Курс JPY/RUB'
    sheet['F1'] = 'Время JPY/RUB'

    sheet['G1'] = 'Результат'

def fill_data(sheet, usd_data, jpy_data):
    for row, (usd, jpy) in enumerate(zip(usd_data, jpy_data)):
        
        sheet[f'A{row + 2}'] = usd['date']
        sheet[f'B{row + 2}'] = usd['rate']
        sheet[f'C{row + 2}'] = usd['time']
        sheet[f'D{row + 2}'] = jpy['date']
        sheet[f'E{row + 2}'] = jpy['rate']
        sheet[f'F{row + 2}'] = jpy['time']
        sheet[f'G{row + 2}'] = usd['rate'] / jpy['rate']

    return row + 2

def apply_format(sheet, max_row_num):
    alignment = Alignment(
                    horizontal='justify',
                    vertical='center',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)
    

    for row in sheet[f"A1:G{max_row_num}"]:
        for cell in row:
            cell.alignment = alignment

    number_format = '# ###0.00 [$₽-419]'

    for row in sheet[f"B2:B{max_row_num}"]:
        for cell in row:
            cell.style = 'Currency'
            cell.number_format = number_format

    for row in sheet[f"E2:E{max_row_num}"]:
        for cell in row:
            cell.style = 'Currency'
            cell.number_format = number_format

    column_widths = []
    for row in sheet[f"A1:G{max_row_num}"]:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value)) > column_widths[i]:
                    column_widths[i] = len(str(cell.value))
            else:
                column_widths += [len(str(cell.value))]
    for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
        sheet.column_dimensions[get_column_letter(i)].width = column_width


if __name__ == '__main__':
    main()